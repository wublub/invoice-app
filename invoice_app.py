# -*- coding: utf-8 -*-
"""发票识别工具 - tkinter 桌面版

可用 PyInstaller 打包成 exe：
  pip install pyinstaller
  pyinstaller --onefile --windowed --name 发票识别工具 invoice_app.py
"""

from __future__ import annotations

import io
import json
import os
import re
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from pathlib import Path

import fitz  # PyMuPDF
import pandas as pd

# ---- 解决打包后路径问题 ----
def get_base_dir() -> Path:
    if getattr(sys, "frozen", False):
        return Path(sys.executable).resolve().parent
    return Path(__file__).resolve().parent

BASE_DIR = get_base_dir()
sys.path.insert(0, str(BASE_DIR))

import invoice_recognizer as ir

CONFIG_PATH = BASE_DIR / ".invoice_ui_config.json"

# ==================== 配置读写 ====================

def load_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            return {}
    return {}


def save_config(cfg: dict) -> None:
    try:
        CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    except Exception:
        pass


# ==================== 工具函数 ====================

def extract_text_from_pdf_bytes(data: bytes) -> str:
    doc = fitz.open(stream=data, filetype="pdf")
    parts: list[str] = []
    for i in range(doc.page_count):
        parts.append(doc.load_page(i).get_text("text"))
    return "\n".join(parts)


def format_date(date_str: str, fmt: str) -> str:
    s = (date_str or "").strip()
    if not re.fullmatch(r"\d{8}", s):
        return s
    y, m, d = s[0:4], s[4:6], s[6:8]
    formats = {
        "YYYYMMDD": f"{y}{m}{d}",
        "YYYY-MM-DD": f"{y}-{m}-{d}",
        "YYYY/MM/DD": f"{y}/{m}/{d}",
        "YYYY.MM.DD": f"{y}.{m}.{d}",
        "YYYY年MM月DD日": f"{y}年{m}月{d}日",
    }
    return formats.get(fmt, s)


def build_excel_bytes(df: pd.DataFrame) -> bytes:
    buf = io.BytesIO()
    df.to_excel(buf, index=False, engine="openpyxl")
    buf.seek(0)

    from openpyxl import load_workbook
    from openpyxl.styles import Alignment

    wb = load_workbook(buf)
    ws = wb.active
    ws.freeze_panes = "A2"

    header_cells = list(ws[1])
    header_to_letter = {(c.value or ""): c.column_letter for c in header_cells}

    width_map = {
        "文件名": 24, "发票日期": 14, "发票号": 24,
        "金额": 12, "购买详细": 60,
        "购买方名称": 30, "购买方纳税人识别号": 26,
        "销售方名称": 30, "销售方纳税人识别号": 26,
        "账号": 26, "开户银行": 30,
    }
    for name, letter in header_to_letter.items():
        ws.column_dimensions[letter].width = width_map.get(name, 18)

    wrap_letters = set()
    if "购买详细" in header_to_letter:
        wrap_letters.add(header_to_letter["购买详细"])

    for r in range(2, ws.max_row + 1):
        for cell in ws[r]:
            cell.alignment = Alignment(vertical="top")
        for letter in wrap_letters:
            ws[f"{letter}{r}"].alignment = Alignment(wrap_text=True, vertical="top")

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def resolve_output_col(df: pd.DataFrame, col_name: str) -> None:
    if col_name in df.columns:
        return
    if "/" not in col_name:
        return
    parts = [p.strip() for p in col_name.split("/") if p.strip()]
    if len(parts) < 2:
        return
    if not all(p in df.columns for p in parts):
        return
    def merge_row(row, fnames=parts):
        vals = []
        for fn in fnames:
            v = str(row.get(fn, "")).strip()
            if v:
                vals.append(v)
        return "/".join(vals)
    df[col_name] = df.apply(merge_row, axis=1)


# ==================== 主应用 ====================

class InvoiceApp:
    DATE_FORMATS = ["YYYYMMDD", "YYYY-MM-DD", "YYYY/MM/DD", "YYYY.MM.DD", "YYYY年MM月DD日"]

    # 全局字体配置
    FONT_NORMAL = ("Microsoft YaHei", 13)
    FONT_BOLD = ("Microsoft YaHei", 13, "bold")
    FONT_SMALL = ("Microsoft YaHei", 11)
    FONT_SMALL_BOLD = ("Microsoft YaHei", 11, "bold")
    FONT_LIST = ("Microsoft YaHei", 13)

    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("发票识别工具")
        self.root.geometry("1100x750")
        self.root.minsize(900, 600)

        # 设置窗口图标
        icon_path = BASE_DIR / "invoice_icon.ico"
        if icon_path.exists():
            try:
                self.root.iconbitmap(str(icon_path))
            except Exception:
                pass

        self.cfg = load_config()
        self.base_fields = list(ir.COLUMNS) + ["文件名"]
        self.extra_fields: list[str] = self.cfg.get("extra_fields", [])
        self.extra_values: dict[str, str] = self.cfg.get("extra_values", {})
        self.output_rows: list[list[str]] = []
        self._load_output_rows()

        self.pdf_paths: list[Path] = []
        self.result_df: pd.DataFrame | None = None

        self._build_ui()

    def _load_output_rows(self):
        cfg_rows = self.cfg.get("output_rows")
        if not cfg_rows:
            self.output_rows = [[c] for c in self.base_fields]
        else:
            self.output_rows = []
            for row in cfg_rows:
                if isinstance(row, str):
                    self.output_rows.append([row])
                elif isinstance(row, list):
                    cleaned = [str(x) for x in row if str(x).strip()]
                    if cleaned:
                        self.output_rows.append(cleaned)
                else:
                    self.output_rows.append([str(row)])

    def _all_fields(self) -> list[str]:
        seen = set()
        out = []
        for f in self.base_fields + self.extra_fields:
            if f not in seen:
                seen.add(f)
                out.append(f)
        return out

    def _used_fields(self) -> set[str]:
        used = set()
        for row in self.output_rows:
            for f in row:
                used.add(f)
        return used

    def _pool_fields(self) -> list[str]:
        used = self._used_fields()
        return [f for f in self._all_fields() if f not in used]

    # ==================== UI 构建 ====================

    def _build_ui(self):
        # 统一放大 ttk 控件字体
        style = ttk.Style()
        style.configure("TNotebook.Tab", font=("Microsoft YaHei", 13, "bold"), padding=[12, 6])
        style.configure("TLabel", font=self.FONT_NORMAL)
        style.configure("TButton", font=self.FONT_NORMAL)
        style.configure("TRadiobutton", font=self.FONT_NORMAL)
        style.configure("TEntry", font=self.FONT_NORMAL)
        style.configure("TLabelframe.Label", font=self.FONT_BOLD)
        style.configure("Treeview", font=self.FONT_NORMAL, rowheight=28)
        style.configure("Treeview.Heading", font=self.FONT_BOLD)

        notebook = ttk.Notebook(self.root)
        notebook.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)

        # Tab 1: 选择发票
        tab1 = ttk.Frame(notebook)
        notebook.add(tab1, text="  1. 选择发票  ")
        self._build_tab_files(tab1)

        # Tab 2: 输出设置
        tab2 = ttk.Frame(notebook)
        notebook.add(tab2, text="  2. 输出设置  ")
        self._build_tab_settings(tab2)

        # Tab 3: 识别与导出
        tab3 = ttk.Frame(notebook)
        notebook.add(tab3, text="  3. 识别与导出  ")
        self._build_tab_result(tab3)

        # 底部致谢栏
        credit = ttk.Label(
            self.root,
            text="产品经理: wublub  ·  大模型: claude-4.6-opus",
            font=self.FONT_SMALL, foreground="#b2bec3",
        )
        credit.pack(side=tk.BOTTOM, pady=(0, 4))

    # ---- Tab 1: 文件选择 ----
    def _build_tab_files(self, parent):
        frame = ttk.LabelFrame(parent, text="选择 PDF 发票文件", padding=10)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        btn_frame = ttk.Frame(frame)
        btn_frame.pack(fill=tk.X, pady=(0, 10))

        ttk.Button(btn_frame, text="选择文件", command=self._select_files).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="选择文件夹", command=self._select_folder).pack(side=tk.LEFT, padx=(0, 10))
        ttk.Button(btn_frame, text="清空列表", command=self._clear_files).pack(side=tk.LEFT)

        self.file_count_var = tk.StringVar(value="已选 0 份 PDF")
        ttk.Label(btn_frame, textvariable=self.file_count_var).pack(side=tk.RIGHT)

        list_frame = ttk.Frame(frame)
        list_frame.pack(fill=tk.BOTH, expand=True)

        self.file_listbox = tk.Listbox(list_frame, selectmode=tk.EXTENDED, font=self.FONT_LIST)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # 自动加载上次的文件夹
        last_folder = self.cfg.get("folder", "")
        if last_folder and Path(last_folder).is_dir():
            self._load_folder(Path(last_folder))

    def _select_files(self):
        files = filedialog.askopenfilenames(
            title="选择 PDF 文件",
            filetypes=[("PDF 文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if files:
            self.pdf_paths = [Path(f) for f in files]
            self._refresh_file_list()

    def _select_folder(self):
        folder = filedialog.askdirectory(title="选择包含 PDF 的文件夹")
        if folder:
            self._load_folder(Path(folder))
            self.cfg["folder"] = folder
            save_config(self.cfg)

    def _load_folder(self, folder: Path):
        self.pdf_paths = sorted(folder.glob("*.pdf"))
        self._refresh_file_list()

    def _clear_files(self):
        self.pdf_paths = []
        self._refresh_file_list()

    def _refresh_file_list(self):
        self.file_listbox.delete(0, tk.END)
        for p in self.pdf_paths:
            self.file_listbox.insert(tk.END, str(p))
        self.file_count_var.set(f"已选 {len(self.pdf_paths)} 份 PDF")

    # ---- Tab 2: 输出设置 ----
    def _build_tab_settings(self, parent):
        # 日期格式
        date_frame = ttk.LabelFrame(parent, text="日期格式", padding=5)
        date_frame.pack(fill=tk.X, padx=10, pady=(10, 5))

        self.date_fmt_var = tk.StringVar(value=self.cfg.get("date_format", "YYYYMMDD"))
        for fmt in self.DATE_FORMATS:
            ttk.Radiobutton(date_frame, text=fmt, variable=self.date_fmt_var, value=fmt).pack(side=tk.LEFT, padx=8)

        # 输出列设置
        col_frame = ttk.LabelFrame(parent, text="输出列设置（拖拽排序 · 双击移动 · 同一列多字段用 / 合并）", padding=8)
        col_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # 左右布局：输出列 | 按钮 | 可用字段池
        body = ttk.Frame(col_frame)
        body.pack(fill=tk.BOTH, expand=True)
        body.columnconfigure(0, weight=3)
        body.columnconfigure(1, weight=0)
        body.columnconfigure(2, weight=2)
        body.rowconfigure(1, weight=1)

        # ---- 左侧：输出列 ----
        ttk.Label(body, text="📋 输出列（上→下 = Excel 左→右，拖拽排序）",
                  font=self.FONT_SMALL_BOLD).grid(row=0, column=0, sticky="w", pady=(0, 4))

        left_list_frame = ttk.Frame(body)
        left_list_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 4))

        self.col_listbox = tk.Listbox(
            left_list_frame, selectmode=tk.SINGLE, font=self.FONT_LIST,
            bg="#fff5f5", fg="#2d3436", selectbackground="#e17055", selectforeground="#fff",
            relief=tk.FLAT, bd=0, highlightthickness=1, highlightcolor="#e17055",
            highlightbackground="#dfe6e9", activestyle="none",
        )
        col_scroll = ttk.Scrollbar(left_list_frame, orient=tk.VERTICAL, command=self.col_listbox.yview)
        self.col_listbox.configure(yscrollcommand=col_scroll.set)
        self.col_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        col_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 拖拽绑定 - 输出列
        self._drag_data = {"source": None, "index": -1, "dragging": False}
        self.col_listbox.bind("<ButtonPress-1>", self._on_col_drag_start)
        self.col_listbox.bind("<B1-Motion>", self._on_col_drag_motion)
        self.col_listbox.bind("<ButtonRelease-1>", self._on_col_drag_end)
        self.col_listbox.bind("<Double-Button-1>", self._on_col_double_click)

        # 输出列操作按钮行
        left_btn = ttk.Frame(body)
        left_btn.grid(row=2, column=0, sticky="ew", pady=(4, 0), padx=(0, 4))
        ttk.Button(left_btn, text="← 移到池", command=self._col_to_pool, width=10).pack(side=tk.LEFT, padx=2)
        ttk.Button(left_btn, text="合并到上一列", command=self._merge_col_up, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(left_btn, text="删除列", command=self._delete_col, width=8).pack(side=tk.LEFT, padx=2)

        # ---- 中间：箭头按钮 ----
        mid = ttk.Frame(body)
        mid.grid(row=1, column=1, padx=8, sticky="ns")
        mid.rowconfigure(0, weight=1)
        mid.rowconfigure(5, weight=1)

        ttk.Frame(mid).grid(row=0)  # spacer
        ttk.Button(mid, text="  ←  ", command=self._pool_to_col, width=6).grid(row=1, pady=4)
        ttk.Button(mid, text="  →  ", command=self._col_to_pool, width=6).grid(row=2, pady=4)
        ttk.Separator(mid, orient=tk.HORIZONTAL).grid(row=3, sticky="ew", pady=8)
        ttk.Button(mid, text="  ↑  ", command=self._move_col_up, width=6).grid(row=4, pady=4)
        ttk.Button(mid, text="  ↓  ", command=self._move_col_down, width=6).grid(row=5, pady=4)

        # ---- 右侧：可用字段池 ----
        ttk.Label(body, text="🗂 可用字段池（双击添加到输出列）",
                  font=self.FONT_SMALL_BOLD).grid(row=0, column=2, sticky="w", pady=(0, 4))

        right_list_frame = ttk.Frame(body)
        right_list_frame.grid(row=1, column=2, sticky="nsew", padx=(4, 0))

        self.pool_listbox = tk.Listbox(
            right_list_frame, selectmode=tk.SINGLE, font=self.FONT_LIST,
            bg="#f0f3f8", fg="#2d3436", selectbackground="#636e72", selectforeground="#fff",
            relief=tk.FLAT, bd=0, highlightthickness=1, highlightcolor="#636e72",
            highlightbackground="#dfe6e9", activestyle="none",
        )
        pool_scroll = ttk.Scrollbar(right_list_frame, orient=tk.VERTICAL, command=self.pool_listbox.yview)
        self.pool_listbox.configure(yscrollcommand=pool_scroll.set)
        self.pool_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        pool_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 双击池字段 -> 添加到输出列
        self.pool_listbox.bind("<Double-Button-1>", self._on_pool_double_click)

        # 预览条
        self.preview_var = tk.StringVar(value="")
        preview_frame = ttk.Frame(col_frame)
        preview_frame.pack(fill=tk.X, pady=(6, 0))
        ttk.Label(preview_frame, text="预览:", font=self.FONT_SMALL_BOLD).pack(side=tk.LEFT)
        ttk.Label(preview_frame, textvariable=self.preview_var,
                  font=self.FONT_SMALL, foreground="#e17055").pack(side=tk.LEFT, padx=4)

        # 自定义字段区
        custom_frame = ttk.LabelFrame(parent, text="自定义字段（添加后进入可用字段池，可设固定值）", padding=5)
        custom_frame.pack(fill=tk.X, padx=10, pady=(5, 5))

        add_row = ttk.Frame(custom_frame)
        add_row.pack(fill=tk.X, pady=(0, 5))

        self.new_field_var = tk.StringVar()
        ttk.Entry(add_row, textvariable=self.new_field_var, width=20).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(add_row, text="添加字段到可用池", command=self._add_custom_field).pack(side=tk.LEFT, padx=(0, 10))

        self.custom_list_frame = ttk.Frame(custom_frame)
        self.custom_list_frame.pack(fill=tk.X)

        # 保存按钮
        save_frame = ttk.Frame(parent)
        save_frame.pack(fill=tk.X, padx=10, pady=5)
        ttk.Button(save_frame, text="💾 保存设置（下次仍保留）", command=self._save_settings).pack(side=tk.LEFT)

        # 初始化列表
        self._refresh_col_lists()
        self._refresh_custom_fields()

    # ---- 拖拽排序 ----
    def _on_col_drag_start(self, event):
        idx = self.col_listbox.nearest(event.y)
        if idx < 0:
            return
        self._drag_data["source"] = "col"
        self._drag_data["index"] = idx
        self._drag_data["dragging"] = False
        self.col_listbox.selection_clear(0, tk.END)
        self.col_listbox.selection_set(idx)

    def _on_col_drag_motion(self, event):
        if self._drag_data["source"] != "col":
            return
        self._drag_data["dragging"] = True
        target = self.col_listbox.nearest(event.y)
        src = self._drag_data["index"]
        if target < 0 or target == src:
            return
        # 交换位置
        self.output_rows[src], self.output_rows[target] = self.output_rows[target], self.output_rows[src]
        self._drag_data["index"] = target
        self._refresh_col_lists()
        self.col_listbox.selection_set(target)
        self.col_listbox.configure(cursor="fleur")

    def _on_col_drag_end(self, event):
        if self._drag_data.get("dragging"):
            self.col_listbox.configure(cursor="")
        self._drag_data = {"source": None, "index": -1, "dragging": False}

    def _on_col_double_click(self, event):
        """双击输出列项 -> 移到可用字段池"""
        sel = self.col_listbox.curselection()
        if not sel:
            return
        if len(self.output_rows) <= 1:
            messagebox.showwarning("提示", "至少保留一列。")
            return
        self.output_rows.pop(sel[0])
        self._refresh_col_lists()

    def _on_pool_double_click(self, event):
        """双击可用字段池项 -> 添加到输出列"""
        sel = self.pool_listbox.curselection()
        if not sel:
            return
        field = self.pool_listbox.get(sel[0]).strip()
        self.output_rows.append([field])
        self._refresh_col_lists()

    def _refresh_col_lists(self):
        # 刷新输出列
        self.col_listbox.delete(0, tk.END)
        for i, row in enumerate(self.output_rows):
            label = "/".join(row) if len(row) > 1 else row[0]
            self.col_listbox.insert(tk.END, f"  第{i+1}列:  {label}")
            # 交替行颜色
            bg = "#fff5f5" if i % 2 == 0 else "#ffe8e5"
            self.col_listbox.itemconfig(i, bg=bg)

        # 刷新可用字段池
        self.pool_listbox.delete(0, tk.END)
        for i, f in enumerate(self._pool_fields()):
            self.pool_listbox.insert(tk.END, f"  {f}")
            bg = "#f0f3f8" if i % 2 == 0 else "#e4e9f0"
            self.pool_listbox.itemconfig(i, bg=bg)

        # 更新预览条
        if hasattr(self, "preview_var"):
            names = []
            for row in self.output_rows:
                names.append("/".join(row) if len(row) > 1 else row[0])
            self.preview_var.set("  |  ".join(names) if names else "（无）")

    def _move_col_up(self):
        sel = self.col_listbox.curselection()
        if not sel or sel[0] == 0:
            return
        i = sel[0]
        self.output_rows[i], self.output_rows[i-1] = self.output_rows[i-1], self.output_rows[i]
        self._refresh_col_lists()
        self.col_listbox.selection_set(i-1)

    def _move_col_down(self):
        sel = self.col_listbox.curselection()
        if not sel or sel[0] >= len(self.output_rows) - 1:
            return
        i = sel[0]
        self.output_rows[i], self.output_rows[i+1] = self.output_rows[i+1], self.output_rows[i]
        self._refresh_col_lists()
        self.col_listbox.selection_set(i+1)

    def _col_to_pool(self):
        """把选中的输出列移回可用字段池"""
        sel = self.col_listbox.curselection()
        if not sel:
            return
        i = sel[0]
        self.output_rows.pop(i)
        self._refresh_col_lists()

    def _delete_col(self):
        """删除选中的输出列（字段回到池中）"""
        sel = self.col_listbox.curselection()
        if not sel:
            return
        i = sel[0]
        if len(self.output_rows) <= 1:
            messagebox.showwarning("提示", "至少保留一列。")
            return
        self.output_rows.pop(i)
        self._refresh_col_lists()

    def _merge_col_up(self):
        """将选中列合并到上一列（用 / 分隔）"""
        sel = self.col_listbox.curselection()
        if not sel or sel[0] == 0:
            messagebox.showinfo("提示", "请选中第2列及以后的列进行合并。")
            return
        i = sel[0]
        self.output_rows[i-1].extend(self.output_rows[i])
        self.output_rows.pop(i)
        self._refresh_col_lists()
        self.col_listbox.selection_set(i-1)

    def _pool_to_col(self):
        """把选中的池字段添加为新的输出列"""
        sel = self.pool_listbox.curselection()
        if not sel:
            return
        field = self.pool_listbox.get(sel[0]).strip()
        self.output_rows.append([field])
        self._refresh_col_lists()

    def _add_new_col_from_pool(self):
        """把选中的池字段添加为新的输出列（等同于 pool_to_col）"""
        self._pool_to_col()

    def _add_custom_field(self):
        name = self.new_field_var.get().strip()
        if not name:
            return
        all_names = set(self.base_fields) | set(self.extra_fields)
        if name in all_names:
            messagebox.showwarning("提示", f"字段 '{name}' 已存在。")
            return
        self.extra_fields.append(name)
        self.new_field_var.set("")
        self._refresh_col_lists()
        self._refresh_custom_fields()
        self.cfg["extra_fields"] = self.extra_fields
        save_config(self.cfg)

    def _refresh_custom_fields(self):
        for w in self.custom_list_frame.winfo_children():
            w.destroy()

        for idx, ef in enumerate(self.extra_fields):
            row = ttk.Frame(self.custom_list_frame)
            row.pack(fill=tk.X, pady=1)

            ttk.Label(row, text=ef, width=15, foreground="#d63031",
                     font=self.FONT_SMALL_BOLD).pack(side=tk.LEFT, padx=(0, 5))

            ttk.Label(row, text="固定值:").pack(side=tk.LEFT)
            val_var = tk.StringVar(value=self.extra_values.get(ef, ""))
            entry = ttk.Entry(row, textvariable=val_var, width=20)
            entry.pack(side=tk.LEFT, padx=(2, 5))
            entry.bind("<FocusOut>", lambda e, name=ef, var=val_var: self._on_custom_val_change(name, var))
            entry.bind("<Return>", lambda e, name=ef, var=val_var: self._on_custom_val_change(name, var))

            ttk.Button(row, text="✖", width=3,
                      command=lambda i=idx: self._delete_custom_field(i)).pack(side=tk.LEFT)

    def _on_custom_val_change(self, name: str, var: tk.StringVar):
        self.extra_values[name] = var.get()
        self.cfg["extra_values"] = self.extra_values
        save_config(self.cfg)

    def _delete_custom_field(self, idx: int):
        removed = self.extra_fields.pop(idx)
        self.extra_values.pop(removed, None)
        # 从输出列中也移除
        new_rows = []
        for row in self.output_rows:
            cleaned = [x for x in row if x != removed]
            if cleaned:
                new_rows.append(cleaned)
        self.output_rows = new_rows if new_rows else [[self.base_fields[0]]]
        self.cfg["extra_fields"] = self.extra_fields
        self.cfg["extra_values"] = self.extra_values
        save_config(self.cfg)
        self._refresh_col_lists()
        self._refresh_custom_fields()

    def _save_settings(self):
        self.cfg["output_rows"] = self.output_rows
        self.cfg["date_format"] = self.date_fmt_var.get()
        self.cfg["extra_fields"] = self.extra_fields
        self.cfg["extra_values"] = self.extra_values
        save_config(self.cfg)
        messagebox.showinfo("提示", "设置已保存。")

    # ---- Tab 3: 识别与导出 ----
    def _build_tab_result(self, parent):
        top = ttk.Frame(parent)
        top.pack(fill=tk.X, padx=10, pady=10)

        ttk.Button(top, text="开始识别（解析 PDF）", command=self._run_recognize).pack(side=tk.LEFT, padx=(0, 10))

        self.progress_var = tk.StringVar(value="")
        ttk.Label(top, textvariable=self.progress_var).pack(side=tk.LEFT)

        # 结果表格
        table_frame = ttk.Frame(parent)
        table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 5))

        # 用 Treeview 作为表格
        self.tree = ttk.Treeview(table_frame, show="headings", selectmode="browse")

        vsb = ttk.Scrollbar(table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)

        # 导出区
        export_frame = ttk.Frame(parent)
        export_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

        ttk.Label(export_frame, text="导出文件名:").pack(side=tk.LEFT)
        self.export_name_var = tk.StringVar(value=self.cfg.get("export_name", "发票识别结果.xlsx"))
        ttk.Entry(export_frame, textvariable=self.export_name_var, width=30).pack(side=tk.LEFT, padx=5)
        ttk.Button(export_frame, text="导出 Excel", command=self._export_excel).pack(side=tk.LEFT, padx=5)

    def _run_recognize(self):
        if not self.pdf_paths:
            messagebox.showwarning("提示", "请先在第1步选择 PDF 文件。")
            return

        self.progress_var.set("识别中...")
        self.root.update()

        # 在线程中执行识别，避免界面卡住
        threading.Thread(target=self._do_recognize, daemon=True).start()

    def _do_recognize(self):
        rows: list[dict] = []
        errors: list[str] = []
        total = len(self.pdf_paths)

        for i, f in enumerate(self.pdf_paths):
            try:
                text = ir.normalize_text(ir.extract_text_from_pdf(f))
                if len(text.strip()) < 30:
                    errors.append(f"{f.name}：提取文本很少，可能是扫描件。")
                r = ir.parse_invoice_from_text(text)
                r["文件名"] = f.name
                rows.append(r)
            except Exception as e:
                errors.append(f"{f.name}：{e}")

            self.root.after(0, lambda v=f"识别中... ({i+1}/{total})": self.progress_var.set(v))

        # 回到主线程更新 UI
        self.root.after(0, lambda: self._on_recognize_done(rows, errors))

    def _on_recognize_done(self, rows: list[dict], errors: list[str]):
        if errors:
            messagebox.showwarning("识别提示", "\n".join(errors))

        if not rows:
            self.progress_var.set("无结果")
            return

        # 构建 DataFrame
        df = pd.DataFrame(rows).fillna("")

        # 自定义字段写入
        for ef in self.extra_fields:
            val = self.extra_values.get(ef, "")
            df[ef] = val

        # 日期格式
        date_fmt = self.date_fmt_var.get()
        if "发票日期" in df.columns:
            df["发票日期"] = df["发票日期"].astype(str).map(lambda x: format_date(x, date_fmt))

        # 发票号/日期当文本
        for col in ["发票号", "发票日期"]:
            if col in df.columns:
                df[col] = df[col].astype(str)

        # 构建输出列
        output_cols = []
        for row in self.output_rows:
            col_name = "/".join(row) if len(row) > 1 else row[0]
            resolve_output_col(df, col_name)
            if col_name not in df.columns:
                df[col_name] = ""
            output_cols.append(col_name)

        self.result_df = df[output_cols]
        self._show_result_table()
        self.progress_var.set(f"识别完成：{len(rows)} 份")

        # 保存设置
        self.cfg["output_rows"] = self.output_rows
        self.cfg["date_format"] = date_fmt
        self.cfg["extra_fields"] = self.extra_fields
        self.cfg["extra_values"] = self.extra_values
        save_config(self.cfg)

    def _show_result_table(self):
        if self.result_df is None:
            return

        # 清空旧数据
        self.tree.delete(*self.tree.get_children())

        cols = list(self.result_df.columns)
        self.tree["columns"] = cols

        for col in cols:
            self.tree.heading(col, text=col)
            # 根据字段名设置列宽
            w = 120
            if col in ("购买详细",):
                w = 300
            elif col in ("发票号", "购买方纳税人识别号", "销售方纳税人识别号", "账号"):
                w = 180
            elif col in ("购买方名称", "销售方名称", "开户银行"):
                w = 200
            elif col in ("金额", "发票日期"):
                w = 100
            self.tree.column(col, width=w, minwidth=60)

        for _, row in self.result_df.iterrows():
            values = []
            for col in cols:
                v = str(row[col])
                # Treeview 不支持换行，用空格替代
                v = v.replace("\n", " | ")
                values.append(v)
            self.tree.insert("", tk.END, values=values)

    def _export_excel(self):
        if self.result_df is None or self.result_df.empty:
            messagebox.showwarning("提示", "请先识别 PDF。")
            return

        name = self.export_name_var.get().strip()
        if not name:
            name = "发票识别结果.xlsx"
        if not name.lower().endswith(".xlsx"):
            name += ".xlsx"

        save_path = filedialog.asksaveasfilename(
            title="保存 Excel",
            defaultextension=".xlsx",
            initialfile=name,
            filetypes=[("Excel 文件", "*.xlsx")]
        )
        if not save_path:
            return

        try:
            excel_bytes = build_excel_bytes(self.result_df)
            Path(save_path).write_bytes(excel_bytes)
            self.cfg["export_name"] = name
            save_config(self.cfg)
            messagebox.showinfo("成功", f"已导出：{save_path}")
        except Exception as e:
            messagebox.showerror("导出失败", str(e))


# ==================== 入口 ====================

def main():
    root = tk.Tk()
    # 设置 DPI 感知（Windows 高分屏）
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)
    except Exception:
        pass
    app = InvoiceApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()

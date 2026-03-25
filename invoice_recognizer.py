# -*- coding: utf-8 -*-
"""发票识别工具（快速原型）

功能：
- 读取当前目录（或指定目录）下的 PDF 发票
- 从 PDF 文本中提取：发票日期、发票号、金额、购买详细、购买方/销售方名称及纳税人识别号、账号、开户银行
- 按固定列格式导出为 Excel（.xlsx）

说明：
- 该版本优先处理"可复制文本"的电子发票 PDF。
- 若是纯扫描图片型 PDF，需要 OCR（后续可加 PaddleOCR / Tesseract）。
"""

from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path

import fitz  # PyMuPDF
import pandas as pd


COLUMNS = [
    "发票日期",
    "发票号",
    "金额",
    "购买详细",
    "购买方名称",
    "购买方纳税人识别号",
    "销售方名称",
    "销售方纳税人识别号",
    "账号",
    "开户银行",
]


def _stdout_utf8() -> None:
    """避免 Windows 终端因默认编码导致打印失败。"""
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass


def extract_text_from_pdf(pdf_path: Path) -> str:
    doc = fitz.open(str(pdf_path))
    parts: list[str] = []
    for i in range(doc.page_count):
        page = doc.load_page(i)
        parts.append(page.get_text("text"))
    return "\n".join(parts)


def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t\u00a0]+", " ", text)
    text = "\n".join([ln.strip() for ln in text.split("\n")])
    return text


def _lines(text: str) -> list[str]:
    return [ln.strip() for ln in text.split("\n") if ln.strip()]


def _find_first_index(lines: list[str], predicate) -> int | None:
    for i, ln in enumerate(lines):
        if predicate(ln):
            return i
    return None


def _is_tax_id_line(ln: str) -> bool:
    return bool(re.fullmatch(r"[A-Za-z0-9]{15,20}", ln))


def _is_company_name(line: str) -> bool:
    return bool(re.search(r"(有限责任公司|有限公司|公司)$", line))


def _is_buyer_like_name(line: str) -> bool:
    return bool(re.search(r"(大学|学院|研究院|学校|医院|研究所|中心)$", line))


def _header_value_lines(lines: list[str]) -> list[str]:
    start = _find_first_index(lines, lambda ln: bool(re.match(r"\d{4}\s*年", ln) or re.match(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", ln)))
    if start is None:
        start = _find_first_index(lines, lambda ln: bool(re.fullmatch(r"\d{8,20}", ln)))
    if start is None:
        return []

    end = _find_first_index(
        lines,
        lambda ln: ln == "项目名称" or ln.startswith("*") or ln.startswith("销方开户银行") or ln.startswith("银行账号"),
    )
    if end is None or end <= start:
        end = len(lines)
    return lines[start + 1 : end]


def _invoice_no_candidates(text: str) -> list[tuple[int, str, int]]:
    candidates: list[tuple[int, str, int]] = []
    seen: set[tuple[int, str]] = set()
    for m in re.finditer(r"\b\d{8,20}\b", text):
        num = m.group(0)
        score = 0
        start = m.start()
        left = text[max(0, start - 40):start]
        around = text[max(0, start - 80): min(len(text), m.end() + 80)]
        if "发票号码" in left or "发票号码" in around:
            score += 120
        if "开票日期" in around or "电子发票" in around:
            score += 40
        if any(tag in around for tag in ("统一社会信用代码", "纳税人识别号")):
            score -= 120
        if len(num) in (12, 14, 15, 16):
            score += 15
        if len(num) >= 18:
            score -= 10
        if num.startswith("0"):
            score -= 5
        key = (start, num)
        if key not in seen:
            seen.add(key)
            candidates.append((score, num, start))
    return candidates


def parse_invoice_no(text: str) -> str:
    m = re.search(r"发票号码\s*[:：]?\s*([0-9]{8,20})", text)
    if m:
        return m.group(1)

    for m in re.finditer(r"发票号码\s*[:：]?", text):
        window = text[m.end() : m.end() + 200]
        date_match = re.search(r"\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日", window)
        if date_match:
            before_date_lines = _lines(window[: date_match.start()])
            nums = [ln for ln in before_date_lines if re.fullmatch(r"\d{8,20}", ln)]
            if nums:
                nums.sort(key=len, reverse=True)
                return nums[0]
            continue
        window_lines = _lines(window)
        nums = [ln for ln in window_lines if re.fullmatch(r"\d{8,20}", ln)]
        if nums:
            nums.sort(key=len, reverse=True)
            return nums[0]

    lines = _lines(text)
    date_idx = _find_first_index(lines, lambda ln: bool(re.match(r"\d{4}\s*年", ln) or re.match(r"\d{4}[-/.]\d{1,2}[-/.]\d{1,2}", ln)))
    item_idx = _find_first_index(lines, lambda ln: ln == "项目名称" or ln.startswith("*"))
    if date_idx is not None:
        start = max(0, date_idx - 3)
        end = item_idx if item_idx is not None and item_idx > start else min(len(lines), date_idx + 3)
        window_lines = lines[start:end]
        nums = [ln for ln in window_lines if re.fullmatch(r"\d{8,20}", ln)]
        if nums:
            nums.sort(key=lambda x: (-len(x), abs(window_lines.index(x) - (date_idx - start))))
            return nums[0]

    candidates = _invoice_no_candidates(text)
    if candidates:
        candidates.sort(key=lambda x: (-x[0], x[2], len(x[1])))
        return candidates[0][1]

    nums = re.findall(r"\b\d{8,20}\b", text)
    if nums:
        nums.sort(key=len, reverse=True)
        return nums[0]
    return ""


def parse_invoice_date(text: str) -> str:
    m = re.search(r"(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        d = int(m.group(3))
        return f"{y:04d}{mo:02d}{d:02d}"

    m = re.search(r"\b(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})\b", text)
    if m:
        y = int(m.group(1))
        mo = int(m.group(2))
        d = int(m.group(3))
        return f"{y:04d}{mo:02d}{d:02d}"

    return ""


def parse_total_amount(text: str) -> float | None:
    """解析价税合计（小写）。

    兼容一些 PDF 提取文本的常见问题：
    - "¥"符号和金额被拆到不同行（例如："¥\n0.43" 或 "14.60\n¥"）
    - 有的票"（小写）"与金额距离较远

    策略：
    1) 优先在"（小写）"附近找金额
    2) 再在"价税合计"附近找最后一个带币种符号的金额
    3) 兜底：全文找最后一个带币种符号的金额
    """

    def to_float(num: str) -> float | None:
        try:
            return float(num.replace(",", ""))
        except Exception:
            return None

    prefix_pat = re.compile(r"[¥￥]\s*([+-]?\d{1,15}(?:,\d{3})*(?:\.\d{2}))")
    suffix_pat = re.compile(r"([+-]?\d{1,15}(?:,\d{3})*(?:\.\d{2}))\s*[¥￥]")

    def find_currency_numbers(s: str) -> list[tuple[int, float]]:
        out: list[tuple[int, float]] = []
        for m in prefix_pat.finditer(s):
            v = to_float(m.group(1))
            if v is not None:
                out.append((m.start(), v))
        for m in suffix_pat.finditer(s):
            v = to_float(m.group(1))
            if v is not None:
                out.append((m.start(), v))
        out.sort(key=lambda x: x[0])
        return out

    i_small = text.rfind("（小写）")
    if i_small != -1:
        window = text[i_small : i_small + 3000]
        cands = find_currency_numbers(window)
        if cands:
            return cands[-1][1]
        nums = re.findall(r"[+-]?\d{1,15}(?:,\d{3})*(?:\.\d{2})", window)
        if nums:
            v = to_float(nums[-1])
            if v is not None:
                return v

    i_total = text.rfind("价税合计")
    if i_total != -1:
        window = text[i_total : i_total + 2500]
        cands = find_currency_numbers(window)
        if cands:
            return cands[-1][1]

    cands = find_currency_numbers(text)
    if cands:
        return cands[-1][1]

    return None


def parse_bank_account(text: str) -> str:
    m = re.search(r"(?:银行账号|账号)\s*[:：]\s*([0-9]{8,30})", text)
    return m.group(1) if m else ""


def parse_bank_name(text: str) -> str:
    m = re.search(r"(?:销方)?开户银行\s*[:：]\s*(.+?)(?:[;；\n]|银行账号|$)", text)
    if m:
        bank = m.group(1).strip().rstrip(";； ")
        if bank:
            return bank
    return ""


def _assign_parties(name_candidates: list[str], tax_candidates: list[str]) -> dict:
    result = {
        "buyer_name": "",
        "buyer_tax_id": "",
        "seller_name": "",
        "seller_tax_id": "",
    }

    unique_names: list[str] = []
    for name in name_candidates:
        if name and name not in unique_names:
            unique_names.append(name)

    buyer_name = next((x for x in unique_names if _is_buyer_like_name(x)), "")
    seller_name = next((x for x in unique_names if _is_company_name(x) and x != buyer_name), "")

    remaining = [x for x in unique_names if x not in {buyer_name, seller_name}]
    if not buyer_name and remaining:
        buyer_name = remaining.pop(0)
    if not seller_name and remaining:
        seller_name = remaining.pop(0)
    if not seller_name and unique_names:
        for x in unique_names:
            if x != buyer_name:
                seller_name = x
                break

    result["buyer_name"] = buyer_name
    result["seller_name"] = seller_name

    unique_taxes: list[str] = []
    for tax in tax_candidates:
        if tax and tax not in unique_taxes:
            unique_taxes.append(tax)

    if buyer_name and seller_name and len(unique_taxes) >= 2:
        if _is_company_name(buyer_name) and _is_buyer_like_name(seller_name):
            buyer_name, seller_name = seller_name, buyer_name
            result["buyer_name"] = buyer_name
            result["seller_name"] = seller_name
        result["buyer_tax_id"] = unique_taxes[0]
        result["seller_tax_id"] = unique_taxes[1]
    elif len(unique_taxes) == 1:
        if buyer_name and not seller_name:
            result["buyer_tax_id"] = unique_taxes[0]
        elif seller_name and not buyer_name:
            result["seller_tax_id"] = unique_taxes[0]
        else:
            result["buyer_tax_id"] = unique_taxes[0]

    return result


def _parse_buyer_seller_block(text: str) -> dict:
    result = {
        "buyer_name": "",
        "buyer_tax_id": "",
        "seller_name": "",
        "seller_tax_id": "",
    }

    lines = _lines(text)
    name_pat = re.compile(r"^名\s*称\s*[:：]\s*(.+)$")
    tax_pat = re.compile(r"^(?:统一社会信用代码[/／]?纳税人识别号|纳税人识别号)\s*[:：]\s*([A-Za-z0-9]{15,20})$")

    names_with_value: list[str] = []
    tax_ids_with_value: list[str] = []
    for ln in lines:
        m = name_pat.match(ln)
        if m:
            val = m.group(1).strip()
            if val and not re.match(r"(项目名称|规格型号|单\s*位)", val):
                names_with_value.append(val)
            continue
        m = tax_pat.match(ln)
        if m:
            tax_ids_with_value.append(m.group(1).strip())

    layout_a = _assign_parties(names_with_value, tax_ids_with_value)
    if layout_a["buyer_name"] and layout_a["seller_name"]:
        result.update(layout_a)

    header_values = _header_value_lines(lines)
    if header_values:
        name_candidates: list[str] = []
        tax_candidates: list[str] = []
        for ln in header_values:
            if _is_tax_id_line(ln):
                tax_candidates.append(ln)
                continue
            if re.fullmatch(r"\d+(?:\.\d+)?", ln):
                continue
            if re.search(r"[¥￥]", ln):
                continue
            if re.match(r"\d{4}\s*年", ln):
                continue
            if re.search(r"[\u4e00-\u9fff]", ln) and len(ln) >= 2:
                name_candidates.append(ln)

        layout_b = _assign_parties(name_candidates, tax_candidates)
        for k, v in layout_b.items():
            if v and not result[k]:
                result[k] = v

    if result["buyer_name"] and result["seller_name"]:
        if _is_company_name(result["buyer_name"]) and _is_buyer_like_name(result["seller_name"]):
            result["buyer_name"], result["seller_name"] = result["seller_name"], result["buyer_name"]
            result["buyer_tax_id"], result["seller_tax_id"] = result["seller_tax_id"], result["buyer_tax_id"]

    return result


def parse_buyer_company(text: str) -> str:
    return _parse_buyer_seller_block(text)["buyer_name"]


def parse_buyer_tax_id(text: str) -> str:
    return _parse_buyer_seller_block(text)["buyer_tax_id"]


def parse_seller_company(text: str) -> str:
    name = _parse_buyer_seller_block(text)["seller_name"]
    if name:
        return name

    idx = text.find("销方开户银行")
    if idx != -1:
        window = text[max(0, idx - 800): idx]
        m_all = re.findall(
            r"[\u4e00-\u9fffA-Za-z0-9（）()·]{4,60}(?:有限责任公司|有限公司)",
            window,
        )
        if m_all:
            return m_all[-1]

    m_all = re.findall(
        r"[\u4e00-\u9fffA-Za-z0-9（）()·]{4,60}(?:有限责任公司|有限公司)",
        text,
    )
    if m_all:
        return m_all[0]

    return ""


def parse_seller_tax_id(text: str) -> str:
    return _parse_buyer_seller_block(text)["seller_tax_id"]


def _extract_item_section(lines: list[str]) -> list[str]:
    category_only_re = re.compile(r"^\*[^*]{1,40}\*$")
    start = _find_first_index(lines, lambda ln: ln == "项目名称")
    if start is None:
        start = _find_first_index(lines, lambda ln: _is_item_start_line(ln) or bool(category_only_re.match(ln)))
        if start is None:
            return []
        raw_section = lines[start:]
    else:
        raw_section = lines[start + 1 :]
        first_item = _find_first_index(raw_section, lambda ln: _is_item_start_line(ln) or bool(category_only_re.match(ln)))
        if first_item is None:
            return []
        raw_section = raw_section[first_item:]

    out: list[str] = []
    for ln in raw_section:
        if (
            ln in ("合", "计")
            or "价税合计" in ln
            or ln.startswith("销方开户银行")
            or ln.startswith("银行账号")
            or ln.startswith("下载次数")
            or ln.startswith("订单号")
            or ln.startswith("开票人")
            or (ln.startswith("¥") and out)
            or re.fullmatch(r"[壹贰叁肆伍陆柒捌玖拾佰仟万亿元角分整零〇]+", ln)
        ):
            break
        out.append(ln)
    return out


def _is_rate_line(ln: str) -> bool:
    return bool(re.fullmatch(r"\d+(?:\.\d+)?%", ln))


def _is_amount_line(ln: str) -> bool:
    return bool(re.fullmatch(r"-?\d+(?:\.\d+)?", ln))


def _is_unit_line(ln: str) -> bool:
    return bool(re.fullmatch(r"[\u4e00-\u9fff]{1,4}", ln)) and ln not in {"合", "计"}


def _is_spec_line(ln: str) -> bool:
    if " " in ln:
        return False
    if not re.search(r"\d", ln):
        return False
    if re.fullmatch(r"-?\d+(?:\.\d+)?", ln):
        return False
    if _is_rate_line(ln) or _is_unit_line(ln):
        return False
    return bool(re.fullmatch(r"[A-Za-z0-9._\-一-龥]+", ln))


def _is_item_start_line(ln: str) -> bool:
    return bool(re.match(r"^\*[\u4e00-\u9fffA-Za-z][^*]{0,39}\*.+", ln))


def _normalize_item_key(s: str) -> str:
    return re.sub(r"[\s，,]+", "", s)


def _join_name_parts(parts: list[str]) -> str:
    out = ""
    for part in parts:
        part = part.strip()
        if not part:
            continue
        if not out:
            out = part
            continue
        prev = out[-1]
        curr = part[0]
        if re.match(r"[A-Za-z0-9]", prev) and re.match(r"[A-Za-z0-9]", curr):
            out += " " + part
        elif prev in {",", "，"}:
            out += part
        elif re.match(r"[A-Za-z0-9]", prev) and re.match(r"[\u4e00-\u9fff]", curr):
            out += part
        elif re.match(r"[\u4e00-\u9fff]", prev) and re.match(r"[A-Za-z0-9*]", curr):
            out += part if curr == "*" else " " + part
        else:
            out += part
    return out.strip()


def _build_item_from_block(block: list[str]) -> str:
    if not block:
        return ""

    first = block[0]
    m = re.match(r"^(\*[^*]{1,40}\*)(.*)$", first)
    if not m and re.fullmatch(r"\*[^*]{1,40}\*", first):
        category = first
        head_name = ""
    else:
        if not m:
            return ""
        category = m.group(1)
        head_name = m.group(2).strip()
    tail = [x for x in block[1:] if x]

    qty = ""
    unit = ""
    spec = ""
    spec_idx = -1

    for idx in range(len(tail) - 1, -1, -1):
        token = tail[idx]
        if not qty and re.fullmatch(r"\d{1,6}", token):
            qty = token
            continue
        if not unit and _is_unit_line(token):
            unit = token
            continue
        if not spec and _is_spec_line(token):
            spec = token
            spec_idx = idx
            continue

    name_candidates = tail
    if spec_idx != -1:
        name_candidates = tail[:spec_idx]

    name_parts: list[str] = []
    if head_name:
        name_parts.append(head_name)

    for token in name_candidates:
        if token.startswith("¥"):
            continue
        if token == unit and unit:
            continue
        if _is_rate_line(token):
            continue
        if _is_amount_line(token):
            continue
        if re.fullmatch(r"[壹贰叁肆伍陆柒捌玖拾佰仟万亿元角分整零〇]+", token):
            continue
        name_parts.append(token)

    name = _join_name_parts(name_parts).strip("，, ")
    parts = [f"{category}{name}".strip()]
    if spec:
        parts.append(spec)
    if qty:
        parts.append(f"{qty}{unit}" if unit else qty)
    return "，".join([p for p in parts if p])


def parse_items(text: str) -> str:
    lines = _lines(text)
    section = _extract_item_section(lines)
    if not section:
        return ""

    blocks: list[list[str]] = []
    current: list[str] = []
    current_category = ""
    category_only_re = re.compile(r"^\*[^*]{1,40}\*$")

    for ln in section:
        if _is_item_start_line(ln):
            if current:
                blocks.append(current)
            current = [ln]
            current_category = ""
            continue
        if category_only_re.match(ln):
            if current:
                blocks.append(current)
            current = [ln]
            current_category = ln
            continue
        if current_category and not current:
            current = [current_category]
        if current:
            current.append(ln)
    if current:
        blocks.append(current)

    items_out: list[str] = []
    seen: set[str] = set()
    for block in blocks:
        item = _build_item_from_block(block)
        if not item:
            continue
        key = _normalize_item_key(item)
        category_match = re.match(r"^(\*[^*]{1,40}\*)", item)
        category = category_match.group(1) if category_match else ""

        replaced = False
        for idx, old_item in enumerate(list(items_out)):
            old_key = _normalize_item_key(old_item)
            old_category_match = re.match(r"^(\*[^*]{1,40}\*)", old_item)
            old_category = old_category_match.group(1) if old_category_match else ""
            same_item = key in old_key or old_key in key
            if category and old_category == category and same_item:
                if len(key) > len(old_key):
                    items_out[idx] = item
                    seen.discard(old_key)
                    seen.add(key)
                replaced = True
                break

        if replaced:
            continue
        if key in seen:
            continue
        seen.add(key)
        items_out.append(item)

    if items_out:
        return "\n".join(items_out)

    inline_re = re.compile(r"^\*[^*]{1,40}\*.+")
    items_inline: list[str] = []
    for ln in lines:
        if inline_re.match(ln):
            items_inline.append(ln)
    return "\n".join(items_inline)


def normalize_amount_for_excel(x: float | None):
    if x is None:
        return ""
    if abs(x - round(x)) < 1e-9:
        return int(round(x))
    return x


def parse_invoice_from_text(text: str) -> dict:
    invoice_no = parse_invoice_no(text)
    invoice_date = parse_invoice_date(text)
    amount = normalize_amount_for_excel(parse_total_amount(text))

    account = parse_bank_account(text)
    seller = parse_seller_company(text)
    seller_tax = parse_seller_tax_id(text)
    bank = parse_bank_name(text)
    buyer = parse_buyer_company(text)
    buyer_tax = parse_buyer_tax_id(text)

    items = parse_items(text)

    return {
        "发票日期": invoice_date,
        "发票号": invoice_no,
        "金额": amount,
        "购买详细": items,
        "购买方名称": buyer,
        "购买方纳税人识别号": buyer_tax,
        "销售方名称": seller,
        "销售方纳税人识别号": seller_tax,
        "账号": account,
        "开户银行": bank,
    }


def find_invoice_files(input_dir: Path) -> list[Path]:
    pdfs = sorted(input_dir.glob("*.pdf"))
    return pdfs


def write_excel(rows: list[dict], output_path: Path) -> None:
    df = pd.DataFrame(rows, columns=COLUMNS)
    if "发票号" in df.columns:
        df["发票号"] = df["发票号"].astype(str)
    if "发票日期" in df.columns:
        df["发票日期"] = df["发票日期"].astype(str)

    df.to_excel(output_path, index=False, engine="openpyxl")

    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment

        wb = load_workbook(output_path)
        ws = wb.active
        ws.freeze_panes = "A2"

        widths = {
            "A": 12,
            "B": 24,
            "C": 10,
            "D": 60,
            "E": 30,
            "F": 24,
            "G": 30,
            "H": 24,
            "I": 26,
            "J": 30,
        }
        for col, w in widths.items():
            ws.column_dimensions[col].width = w

        for r in range(2, ws.max_row + 1):
            cell = ws[f"D{r}"]
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            for c in ["A", "B", "C", "E", "F", "G", "H", "I", "J"]:
                ws[f"{c}{r}"].alignment = Alignment(vertical="top")

        wb.save(output_path)
    except Exception:
        pass


def main() -> int:
    _stdout_utf8()

    ap = argparse.ArgumentParser(description="发票识别并导出 Excel")
    ap.add_argument("--input", default=".", help="发票文件所在目录（默认当前目录）")
    ap.add_argument("--output", default="发票识别结果.xlsx", help="输出 Excel 文件名")
    args = ap.parse_args()

    input_dir = Path(args.input).resolve()
    output_path = Path(args.output).resolve()

    files = find_invoice_files(input_dir)
    if not files:
        print(f"未找到 PDF 发票文件：{input_dir}")
        return 2

    rows: list[dict] = []
    for f in files:
        try:
            text = normalize_text(extract_text_from_pdf(f))
            row = parse_invoice_from_text(text)
            row["_文件"] = f.name
            rows.append(row)
        except Exception as e:
            print(f"解析失败：{f.name} -> {e}")

    rows_excel = [{k: r.get(k, "") for k in COLUMNS} for r in rows]
    write_excel(rows_excel, output_path)

    print("已导出：", str(output_path))
    print("识别预览：")
    for r in rows_excel:
        print("-" * 60)
        for k in COLUMNS:
            v = r.get(k, "")
            if k == "购买详细" and isinstance(v, str) and "\n" in v:
                print(f"{k}:\n{v}")
            else:
                print(f"{k}: {v}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
